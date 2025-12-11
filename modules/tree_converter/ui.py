"""
Tree Converter Module - UI Layer
Convert Excel tree hierarchies to parent-child format
"""

import streamlit as st
import pandas as pd
import io
from datetime import datetime
import sys
import os

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from shared.workflow import send_to_module

def render(workflow_receiver=None):
    """
    Render the Tree Converter module
    
    Args:
        workflow_receiver: Function to call when sending data to another module
    """
    
    # Import engine here to avoid circular imports
    from modules.tree_converter.engine import TreeParser
    
    st.markdown("### Convert Excel Tree to Parent-Child Format")
    
    st.markdown("""
    <div class="info-box">
        <p><strong>1. Enter your Vena dimension name</strong> (e.g., Account, Cost Center, Products)</p>
        <p><strong>2. Upload your Excel hierarchy tree</strong> with:</p>
        <ul>
            <li><strong>Columns A-J (1-10):</strong> Hierarchy levels (position = level)</li>
            <li><strong>Column K (11):</strong> Member aliases (optional)</li>
            <li><strong>Column L (12):</strong> Operators: + (add), - (subtract), ~ (ignore) - defaults to + if blank</li>
        </ul>
        <p><strong>The converter will:</strong> Parse the tree structure, validate relationships, and export a clean parent-child table ready for Vena import.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize module-specific session state
    if 'tree_results_df' not in st.session_state:
        st.session_state.tree_results_df = None
    if 'tree_results_stats' not in st.session_state:
        st.session_state.tree_results_stats = None
    if 'tree_results_errors' not in st.session_state:
        st.session_state.tree_results_errors = None
    if 'tree_results_warnings' not in st.session_state:
        st.session_state.tree_results_warnings = None
    if 'tree_results_vis' not in st.session_state:
        st.session_state.tree_results_vis = None
    if 'tree_file_uploader_key' not in st.session_state:
        st.session_state.tree_file_uploader_key = 0
    
    # Configuration
    st.markdown("---")
    st.markdown("#### Vena Configuration")
    
    dimension_name = st.text_input(
        "Dimension Name (e.g., Account, Cost Center, Products)",
        value="Account",
        help="Enter your Vena dimension name - this will be used in the output file",
        placeholder="Account",
        key="tree_dim_name"
    )
    
    if not dimension_name or not dimension_name.strip():
        st.warning("Please enter a dimension name")
    
    st.markdown("""
    <div class="tip-box">
        <small>
        • Operators default to <code>+</code> (additive rollup) if Column L is blank<br>
        • Use <code>-</code> for contra accounts (Returns, Discounts)<br>
        • Use <code>~</code> to ignore in rollup
        </small>
    </div>
    """, unsafe_allow_html=True)
    
    operator_default = "+"
    
    # File upload
    st.markdown("---")
    uploaded_file = st.file_uploader(
        "Upload Excel Tree File",
        type=['xlsx', 'xls'],
        help="Upload your Excel file with visual hierarchy structure",
        key=f"tree_file_uploader_{st.session_state.tree_file_uploader_key}"
    )
    
    if uploaded_file is not None:
        if not dimension_name or not dimension_name.strip():
            st.error("Please enter a dimension name before uploading a file")
        else:
            st.success(f"File loaded: {uploaded_file.name}")
            
            # Convert button
            if st.button("Convert Tree", type="primary", width="stretch", key="tree_convert_btn"):
                with open("temp_tree.xlsx", "wb") as f:
                    f.write(uploaded_file.getbuffer())
                
                with st.spinner('Parsing tree structure...'):
                    parser = TreeParser()
                    tree, errors, warnings, stats = parser.parse_excel_tree('temp_tree.xlsx')
                
                st.session_state.tree_results_stats = stats
                st.session_state.tree_results_errors = errors
                st.session_state.tree_results_warnings = warnings
                
                # Use custom visualization if there are duplicate members
                has_duplicates = any(w['type'] == 'Repeated Member' for w in warnings)
                if has_duplicates:
                    st.session_state.tree_results_vis = parser.visualize_hierarchy_with_duplicates()
                else:
                    st.session_state.tree_results_vis = parser.get_tree_visualization()
                
                if stats['errors'] == 0:
                    st.session_state.tree_results_df = parser.tree_to_parent_child(
                        dimension_name=dimension_name.strip(),
                        operator=operator_default
                    )
            
            # Display results
            if st.session_state.tree_results_stats is not None:
                stats = st.session_state.tree_results_stats
                errors = st.session_state.tree_results_errors
                warnings = st.session_state.tree_results_warnings
                
                st.markdown("---")
                
                # KPI Badges - v8 Elegant Style
                badges_html = f'''
                <div style="text-align: center; margin: 20px 0;">
                    <span class="summary-badge" style="background: #e3f2fd; color: #1565c0;">{stats['total_members']} Total Members</span>
                    <span class="summary-badge" style="background: #e3f2fd; color: #1565c0;">{stats['max_depth']} Max Depth</span>
                    <span class="summary-badge" style="background: #e3f2fd; color: #1565c0;">{stats['leaf_count']} Leaf Nodes</span>
                    <span class="summary-badge badge-warning">{stats['warnings']} Warnings</span>
                    <span class="summary-badge badge-error">{stats['errors']} Errors</span>
                </div>
                '''
                st.markdown(badges_html, unsafe_allow_html=True)
                
                st.markdown("---")
                
                # Display errors
                if errors:
                    st.markdown("---")
                    st.markdown("#### Errors Found")
                    for err in errors:
                        st.markdown(f"""
                        <div class="error-box">
                            <strong>Row {err['row']}: {err['type']}</strong><br>
                            Member: {err['member']}<br>
                            {err['message']}
                        </div>
                        """, unsafe_allow_html=True)
                
                # Display warnings
                if warnings:
                    st.markdown("---")
                    st.markdown("#### Warnings")
                    for warn in warnings:
                        rows_str = ', '.join(map(str, warn['rows'])) if warn['rows'] else 'N/A'
                        st.markdown(f"""
                        <div class="warning-box">
                            <strong>{warn['type']}</strong><br>
                            Member: {warn['member']}<br>
                            {warn['message']}<br>
                            <small>Rows: {rows_str}</small>
                        </div>
                        """, unsafe_allow_html=True)
                
                # Success path
                if stats['errors'] == 0:
                    if stats['warnings'] > 0:
                        st.markdown("""
                        <div class="info-box">
                            <strong>Hierarchy Built with Warnings</strong><br>
                            Your hierarchy structure has been processed. Review the warnings above and proceed if acceptable.
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.markdown("""
                        <div class="success-box">
                            <strong>Tree Parsed Successfully</strong><br>
                            Your hierarchy structure is valid and ready to convert.
                        </div>
                        """, unsafe_allow_html=True)
                    
                    st.markdown("---")
                    st.markdown("#### Tree Visualization")
                    st.code(st.session_state.tree_results_vis, language=None)
                    
                    st.markdown("---")
                    st.markdown("#### Vena Hierarchy Table (6-Column Format)")
                    
                    df = st.session_state.tree_results_df
                    st.dataframe(df, width="stretch", height=400)
                    
                    st.markdown("---")
                    st.markdown("#### Download or Validate")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        # Excel download
                        output_excel = io.BytesIO()
                        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name='Sheet1')
                            worksheet = writer.sheets['Sheet1']
                            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                            plain_font = Font(name='Calibri', size=11, bold=False)
                            plain_alignment = Alignment(horizontal='left', vertical='top')
                            no_border = Border(left=Side(style=None), right=Side(style=None), 
                                             top=Side(style=None), bottom=Side(style=None))
                            no_fill = PatternFill(fill_type=None)
                            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                                for cell in row:
                                    cell.font = plain_font
                                    cell.alignment = plain_alignment
                                    cell.border = no_border
                                    cell.fill = no_fill
                        output_excel.seek(0)
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        st.download_button(
                            label="Download Excel",
                            data=output_excel,
                            file_name=f"vena_hierarchy_{timestamp}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="tree_download_excel"
                        )
                    
                    with col2:
                        # CSV download
                        csv = df.to_csv(index=False)
                        st.download_button(
                            label="Download CSV",
                            data=csv,
                            file_name=f"vena_hierarchy_{timestamp}.csv",
                            mime="text/csv",
                            key="tree_download_csv"
                        )
                    
                    with col3:
                        # WORKFLOW: Send to validator
                        if st.button("→ Validate This", type="primary", width="stretch", key="tree_send_to_validator"):
                            send_to_module(df.copy(), 'tree_converter', 'hierarchy_validator')
                            st.success("✓ Data sent to Hierarchy Validator!")
                            st.info("Switch to the Hierarchy Validator tab to continue")
                    
                    
                    # Clear button
                    st.markdown("---")
                    if st.button("Clear Results and Load New File", width="stretch", key="tree_clear"):
                        st.session_state.tree_results_df = None
                        st.session_state.tree_results_stats = None
                        st.session_state.tree_results_errors = None
                        st.session_state.tree_results_warnings = None
                        st.session_state.tree_results_vis = None
                        st.session_state.tree_file_uploader_key += 1
                        st.rerun()
    
    else:
        # Show template download when no file
        st.markdown("---")
        st.markdown("#### Need a Template?")
        st.markdown("""
        <div class="info-box">
            <p>Download our sample template to see the expected format for Excel tree hierarchies:</p>
        </div>
        """, unsafe_allow_html=True)
        
        try:
            # Get path to project root (go up from modules/tree_converter/ui.py to project root)
            current_file = os.path.abspath(__file__)
            tree_converter_dir = os.path.dirname(current_file)  # modules/tree_converter
            modules_dir = os.path.dirname(tree_converter_dir)    # modules
            project_root = os.path.dirname(modules_dir)          # project root
            template_path = os.path.join(project_root, 'assets', 'templates', 'sample_hierarchy_tree.xlsx')
            
            with open(template_path, 'rb') as f:
                st.download_button(
                    label="Download Sample Template",
                    data=f,
                    file_name="sample_hierarchy_tree.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width="stretch",
                    key="tree_template_download"
                )
        except:
            st.info("Sample template will be available after setup")
